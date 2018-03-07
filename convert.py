# https://gist.github.com/riccardobenini/6764495
import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0

    sheet = book.sheet_by_index(0)
    sheet.cell()

    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.active

    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)

    print(book1)

    return book1